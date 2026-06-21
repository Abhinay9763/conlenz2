from __future__ import annotations

import json
import os
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from app.core.rules import Finding


def build_report(
    *,
    scan_type: str,
    folder: str,
    findings: list[Finding],
    scanned_files: int = 0,
    scanned_file_paths: list[str] = None,
    elapsed_seconds: float | None = None,
) -> dict[str, Any]:
    if scanned_file_paths is None:
        scanned_file_paths = []
    flagged_files = len({finding.file_path for finding in findings})
    return {
        "report_version": 1,
        "scan_type": scan_type,
        "folder": folder,
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "files_scanned": scanned_files,
        "files_flagged": flagged_files,
        "elapsed_seconds": elapsed_seconds,
        "scanned_file_paths": scanned_file_paths,
        "findings": [finding.__dict__ for finding in findings],
    }


def write_report_json(report: dict[str, Any], output_dir: Path) -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    path = output_dir / f"scan_report_{timestamp}.json"
    path.write_text(json.dumps(report, indent=2), encoding="utf-8")
    return path


def export_report_excel(report: dict[str, Any], export_path: Path) -> None:
    try:
        import openpyxl
    except Exception as exc:
        raise RuntimeError("openpyxl is required for Excel export") from exc

    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    export_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = openpyxl.Workbook()

    # ── Palette ───────────────────────────────────────────────────────────────
    DARK_BLUE  = "1A3A5C"
    MID_BLUE   = "2E6DA4"
    LIGHT_BLUE = "D6E8F7"
    WHITE      = "FFFFFF"
    GREY_ROW   = "F5F8FC"
    SEV_FILL   = {
        "high":   PatternFill("solid", fgColor="FFD0D0"),
        "medium": PatternFill("solid", fgColor="FFF3CC"),
        "low":    PatternFill("solid", fgColor="D8F5D0"),
    }

    findings      = report.get("findings", [])
    files_scanned = report.get("files_scanned", 0)
    files_flagged = report.get("files_flagged", 0)
    elapsed       = report.get("elapsed_seconds", "—")
    scan_type     = str(report.get("scan_type", "")).capitalize()
    generated_at  = report.get("generated_at", "")

    # ── Sheet 1: Summary ─────────────────────────────────────────────────────
    summary = workbook.active
    summary.title = "Summary"
    summary.column_dimensions["A"].width = 22
    summary.column_dimensions["B"].width = 55

    def _header_cell(sheet, row, col, value, bg=DARK_BLUE, size=11):
        c = sheet.cell(row=row, column=col, value=value)
        c.font = Font(bold=True, color=WHITE, size=size)
        c.fill = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(vertical="center", horizontal="left", indent=1)
        return c

    def _label_cell(sheet, row, col, value):
        c = sheet.cell(row=row, column=col, value=value)
        c.font = Font(bold=True, color=DARK_BLUE)
        c.fill = PatternFill("solid", fgColor=LIGHT_BLUE)
        c.alignment = Alignment(vertical="center", indent=1)
        return c

    def _value_cell(sheet, row, col, value):
        c = sheet.cell(row=row, column=col, value=value)
        c.alignment = Alignment(vertical="center", indent=1)
        return c

    # Title banner
    summary.merge_cells("A1:B1")
    t = summary.cell(row=1, column=1, value="Conlenz Audit Report")
    t.font = Font(bold=True, size=15, color=WHITE)
    t.fill = PatternFill("solid", fgColor=DARK_BLUE)
    t.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    summary.row_dimensions[1].height = 36

    # Metadata rows
    meta = [
        ("Scan Type",     scan_type),
        ("Folder",        report.get("folder", "")),
        ("Generated At",  generated_at),
        ("Elapsed",       f"{elapsed}s" if elapsed != "—" else "—"),
    ]
    for i, (label, val) in enumerate(meta, start=2):
        _label_cell(summary, i, 1, label)
        _value_cell(summary, i, 2, val)
        summary.row_dimensions[i].height = 18

    # Spacer
    summary.append([])

    # Stats section header
    stats_hdr_row = len(meta) + 3
    _header_cell(summary, stats_hdr_row, 1, "Metric", bg=MID_BLUE)
    _header_cell(summary, stats_hdr_row, 2, "Value",  bg=MID_BLUE)
    summary.row_dimensions[stats_hdr_row].height = 20

    sev_counts = {"high": 0, "medium": 0, "low": 0}
    for f in findings:
        sev_counts[str(f.get("severity", "")).lower()] = \
            sev_counts.get(str(f.get("severity", "")).lower(), 0) + 1

    stats = [
        ("Files Scanned",    files_scanned),
        ("Files Flagged",    files_flagged),
        ("Total Findings",   len(findings)),
        ("High Severity",    sev_counts.get("high",   0)),
        ("Medium Severity",  sev_counts.get("medium", 0)),
        ("Low Severity",     sev_counts.get("low",    0)),
    ]
    for r_off, (metric, val) in enumerate(stats, start=stats_hdr_row + 1):
        fill = GREY_ROW if r_off % 2 == 0 else WHITE
        for col, v in enumerate([metric, val], start=1):
            c = summary.cell(row=r_off, column=col, value=v)
            c.fill = PatternFill("solid", fgColor=fill)
            c.alignment = Alignment(vertical="center", indent=1)
            if col == 1:
                c.font = Font(bold=True)
        summary.row_dimensions[r_off].height = 17

    # ── Sheet 2: Findings ─────────────────────────────────────────────────────
    fd = workbook.create_sheet("Findings")

    HEADERS    = ["#", "File",  "Rule",  "Severity",  "Confidence", "Snippet",  "Location"]
    COL_WIDTHS = [ 5,    55,     20,       12,           12,           60,         20]

    for col, (hdr, w) in enumerate(zip(HEADERS, COL_WIDTHS), start=1):
        c = fd.cell(row=1, column=col, value=hdr)
        c.font = Font(bold=True, color=WHITE, size=10)
        c.fill = PatternFill("solid", fgColor=DARK_BLUE)
        c.alignment = Alignment(horizontal="center", vertical="center")
        fd.column_dimensions[get_column_letter(col)].width = w
    fd.row_dimensions[1].height = 22

    for row_idx, item in enumerate(findings, start=2):
        sev    = str(item.get("severity", "")).lower()
        f_fill = SEV_FILL.get(sev)
        row_bg = f_fill or PatternFill("solid", fgColor=(GREY_ROW if row_idx % 2 == 0 else WHITE))
        values = [
            row_idx - 1,
            item.get("file_path",  ""),
            item.get("rule",       ""),
            str(item.get("severity",   "")).capitalize(),
            str(item.get("confidence", "")).capitalize(),
            item.get("snippet",   ""),
            item.get("location",  ""),
        ]
        for col, val in enumerate(values, start=1):
            c = fd.cell(row=row_idx, column=col, value=val)
            c.fill = row_bg
            c.alignment = Alignment(wrap_text=True, vertical="top", indent=1)
        fd.row_dimensions[row_idx].height = 15

    fd.freeze_panes = "A2"
    if findings:
        fd.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{len(findings) + 1}"

    # ── Sheet 3: Scanned Files ────────────────────────────────────────────────
    scanned_paths = report.get("scanned_file_paths", [])
    if scanned_paths:
        sf = workbook.create_sheet("Scanned Files")
        c = sf.cell(row=1, column=1, value="File Path")
        c.font = Font(bold=True, color=WHITE, size=10)
        c.fill = PatternFill("solid", fgColor=DARK_BLUE)
        c.alignment = Alignment(horizontal="left", vertical="center")
        sf.column_dimensions["A"].width = 80
        sf.row_dimensions[1].height = 22

        for row_idx, path_str in enumerate(scanned_paths, start=2):
            fill = GREY_ROW if row_idx % 2 == 0 else WHITE
            cell = sf.cell(row=row_idx, column=1, value=path_str)
            cell.fill = PatternFill("solid", fgColor=fill)
            cell.alignment = Alignment(wrap_text=True, vertical="center", indent=1)
            sf.row_dimensions[row_idx].height = 15

    workbook.save(export_path)


def _load_env_value(key: str, env_path: Path | None = None) -> str | None:
    value = os.environ.get(key)
    if value:
        return value.strip()

    if env_path is None:
        env_path = Path(__file__).resolve().parents[2] / ".env"
    if not env_path.exists():
        return None

    for raw_line in env_path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        name, raw_value = line.split("=", 1)
        if name.strip() != key:
            continue
        return raw_value.strip().strip('"')
    return None


def build_report_email_body(report: dict[str, Any]) -> str:
    return "\n".join(
        [
            "Conlenz Scan Summary",
            "",
            f"Scan type: {report.get('scan_type', '')}",
            f"Folder: {report.get('folder', '')}",
            f"Generated at: {report.get('generated_at', '')}",
            f"Files scanned: {report.get('files_scanned', 0)}",
            f"Files flagged: {report.get('files_flagged', 0)}",
            f"Elapsed seconds: {report.get('elapsed_seconds', '')}",
            "",
            "See attached Excel report for full details.",
        ]
    )


def send_report_email(*, report: dict[str, Any], excel_path: Path, recipient: str) -> None:
    try:
        import resend
    except ImportError as exc:
        raise RuntimeError(
            "resend package is required for email: pip install 'resend>=2.0.0'"
        ) from exc

    api_key = _load_env_value("RESEND_KEY")
    sender  = _load_env_value("RESEND_MAIL")
    if not api_key:
        raise RuntimeError("RESEND_KEY not configured")
    if not sender:
        raise RuntimeError("RESEND_MAIL not configured")

    resend.api_key = api_key

    params: resend.Emails.SendParams = {
        "from":    sender,
        "to":      [recipient],
        "subject": "Conlenz Scan Report",
        "text":    build_report_email_body(report),
        "attachments": [
            {
                "filename": excel_path.name,
                # resend SDK expects content as list[int] (raw bytes)
                "content": list(excel_path.read_bytes()),
            }
        ],
    }

    resend.Emails.send(params)


def write_github_summary(report: dict[str, Any], summary_path: str) -> None:
    """Write a rich Markdown step summary to the GitHub Actions summary file."""
    findings = report.get("findings", [])
    files_scanned = report.get("files_scanned", 0)
    files_flagged = report.get("files_flagged", 0)
    elapsed = report.get("elapsed_seconds", None)
    scan_type = report.get("scan_type", "deep")
    generated_at = report.get("generated_at", "")

    # Parse timestamp for display
    try:
        from datetime import datetime, timezone
        dt = datetime.fromisoformat(generated_at.replace("Z", "+00:00"))
        time_display = dt.strftime("%Y-%m-%d at %H:%M UTC")
    except Exception:
        time_display = generated_at

    elapsed_display = f"{elapsed}s" if elapsed is not None else "—"

    # Count by severity
    severity_counts: dict[str, int] = {}
    for f in findings:
        sev = str(f.get("severity", "unknown")).lower()
        severity_counts[sev] = severity_counts.get(sev, 0) + 1

    _SEV_EMOJI = {"high": "🔴", "medium": "🟡", "low": "🟢"}

    has_high = severity_counts.get("high", 0) > 0
    status_icon = "🚨" if has_high else ("⚠️" if findings else "✅")
    status_text = (
        "Critical Issues Found" if has_high
        else "Issues Found" if findings
        else "All Clear"
    )

    lines: list[str] = []

    # ── Header ────────────────────────────────────────────────────────────────
    lines.append(f"# {status_icon} Conlenz Audit — {status_text}\n")
    lines.append(
        f"> Scanned **{time_display}** &nbsp;·&nbsp; "
        f"Mode: `{scan_type}` &nbsp;·&nbsp; "
        f"Duration: `{elapsed_display}`\n"
    )

    # ── Stats overview ────────────────────────────────────────────────────────
    lines.append("## 📊 Overview\n")
    lines.append("| Metric | Value |")
    lines.append("|--------|------:|")
    lines.append(f"| 📁 Files scanned | **{files_scanned}** |")
    lines.append(f"| 🚩 Files flagged | **{files_flagged}** |")
    lines.append(f"| 📋 Total findings | **{len(findings)}** |")
    lines.append(f"| ⏱️ Elapsed | `{elapsed_display}` |")
    lines.append("")

    # ── Scanned Files (Top 10) ────────────────────────────────────────────────
    scanned_paths = report.get("scanned_file_paths", [])
    if scanned_paths:
        lines.append("## 📁 Scanned Files\n")
        display_paths = scanned_paths[:10]
        for p in display_paths:
            lines.append(f"- `{p}`")
        if len(scanned_paths) > 10:
            lines.append(f"\n_... and {len(scanned_paths) - 10} more files_")
        lines.append("\n---\n")

    if not findings:
        lines.append("---\n")
        lines.append("✨ **No sensitive data detected.** The repository is clean.\n")
    else:
        # ── Severity breakdown ────────────────────────────────────────────────
        lines.append("## 🔎 Severity Breakdown\n")
        lines.append("| Severity | Count |")
        lines.append("|----------|------:|")
        for sev_key in ("high", "medium", "low"):
            count = severity_counts.get(sev_key, 0)
            if count:
                emoji = _SEV_EMOJI.get(sev_key, "⚪")
                lines.append(f"| {emoji} {sev_key.capitalize()} | {count} |")
        lines.append("")

        # ── Findings table (capped at 50 rows) ────────────────────────────────
        display = findings[:50]
        lines.append(f"## 📋 Findings{' (top 50)' if len(findings) > 50 else ''}\n")
        lines.append("| # | File | Rule | Severity | Confidence | Snippet |")
        lines.append("|---|------|------|----------|------------|---------|")
        for idx, item in enumerate(display, 1):
            fp = str(item.get("file_path", ""))
            # Show only the last two path components to keep the table narrow
            parts = Path(fp).parts
            short_path = "/".join(parts[-2:]) if len(parts) >= 2 else fp
            rule = item.get("rule", "—")
            sev = str(item.get("severity", "—")).lower()
            conf = str(item.get("confidence", "—"))
            snippet = str(item.get("snippet", ""))[:80].replace("|", "\\|")
            sev_display = f"{_SEV_EMOJI.get(sev, '⚪')} {sev.capitalize()}"
            lines.append(
                f"| {idx} | `{short_path}` | {rule} | {sev_display} | {conf} | `{snippet}` |"
            )
        lines.append("")

    lines.append("---")
    lines.append(
        "_Generated by [Conlenz](https://github.com/Abhinay9763/conlenz2) "
        "&nbsp;·&nbsp; Full report uploaded as a workflow artifact._\n"
    )

    with open(summary_path, "a", encoding="utf-8") as fh:
        fh.write("\n".join(lines))

